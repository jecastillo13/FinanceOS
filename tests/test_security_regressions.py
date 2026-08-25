from datetime import date, datetime, timedelta
from io import BytesIO, StringIO
import csv
import sqlite3

import pytest
from cryptography.fernet import Fernet
from openpyxl import load_workbook
from pydantic import ValidationError
from PIL import Image
from pypdf import PdfReader, PdfWriter
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker
from sqlalchemy.pool import StaticPool

from api.schemas import CuentaCrear
from core.database import Base
from core.models import SesionUsuario
from core.services import auth_service
from core.services.attachment_service import AttachmentService
from core.services.auth_service import AuthService
from core.services.backup_service import BackupService
from core.services.mfa_service import MfaService
from core.services.report_service import ReportService
from core.services.validation import monto_positivo


def _reporte_con_texto(texto):
    return {
        "anio": 2026, "mes": 8, "ingresos_cop": 1, "gastos_cop": 0,
        "balance_cop": 1, "filas": [{
            "fecha": date(2026, 8, 25), "tipo": "Ingreso", "descripcion": texto,
            "categoria": "Salario", "grupo": "Trabajo", "cuenta": "Banco",
            "moneda": "COP", "valor_original": 1, "valor_cop": 1,
            "observaciones": " @SUM(1+1)",
        }],
    }


@pytest.mark.parametrize("texto", ["=WEBSERVICE(\"https://example.invalid\")", "+1+1", "-1+1", "@SUM(1+1)", "\t=1+1", "  =1+1"])
def test_exportaciones_neutralizan_formulas_sin_convertir_numeros(texto):
    reporte = _reporte_con_texto(texto)
    filas = list(csv.reader(StringIO(ReportService.generar_csv(reporte).decode("utf-8-sig"))))
    assert filas[1][2] == f"'{texto}"

    libro = load_workbook(BytesIO(ReportService.generar_excel(reporte)), data_only=False)
    fila = list(libro["Movimientos"].iter_rows(min_row=2, max_row=2))[0]
    assert fila[2].value == f"'{texto}"
    assert fila[2].data_type == "s"
    assert fila[7].data_type == "n"


def test_valores_no_finitos_y_excesivos_son_rechazados():
    for valor in (float("nan"), float("inf"), -float("inf"), 10**20):
        with pytest.raises((ValueError, ValidationError)):
            CuentaCrear(nombre="Cuenta", tipo="Ahorros", saldo=valor)
        with pytest.raises(ValueError):
            monto_positivo(valor)


def test_imagen_se_decodifica_y_elimina_metadatos():
    salida = BytesIO()
    Image.new("RGB", (64, 64), "white").save(salida, "JPEG", comment=b"dato-privado")
    saneada = AttachmentService.validar_y_sanear(salida.getvalue(), "image/jpeg")
    with Image.open(BytesIO(saneada)) as imagen:
        assert imagen.size == (64, 64)
        assert not imagen.info.get("comment")


def test_pdf_activo_es_rechazado():
    escritor = PdfWriter()
    escritor.add_blank_page(width=100, height=100)
    escritor.add_js("app.alert('no')")
    salida = BytesIO()
    escritor.write(salida)
    with pytest.raises(ValueError, match="contenido activo"):
        AttachmentService.validar_y_sanear(salida.getvalue(), "application/pdf")


def test_pdf_activo_con_nombres_escapados_es_rechazado():
    escritor = PdfWriter()
    escritor.add_blank_page(width=100, height=100)
    escritor.add_js("app.alert('no')")
    salida = BytesIO()
    escritor.write(salida)
    evasivo = salida.getvalue().replace(b"/JavaScript", b"/Java#53cript").replace(b"/JS", b"/J#53")
    with pytest.raises(ValueError, match="contenido activo|dañado"):
        AttachmentService.validar_y_sanear(evasivo, "application/pdf")


def test_pdf_normal_se_reconstruye_sin_contenido_activo():
    escritor = PdfWriter()
    escritor.add_blank_page(width=100, height=100)
    salida = BytesIO()
    escritor.write(salida)
    saneado = AttachmentService.validar_y_sanear(salida.getvalue(), "application/pdf")
    assert len(saneado) > 0
    assert len(PdfReader(BytesIO(saneado)).pages) == 1


def test_respaldo_rechaza_ruta_de_adjunto_fuera_de_uploads(tmp_path):
    database = tmp_path / "respaldo.db"
    conexion = sqlite3.connect(database)
    conexion.execute("CREATE TABLE adjuntos_movimiento (id INTEGER PRIMARY KEY, ruta TEXT, tamano INTEGER)")
    conexion.execute("INSERT INTO adjuntos_movimiento VALUES (1, '../../secreto.txt', 10)")
    conexion.commit()
    conexion.close()
    with pytest.raises(ValueError, match="rutas de adjuntos"):
        BackupService._preparar_adjuntos(database, {})


def test_mfa_fallido_bloquea_y_sesion_inactiva_expira(monkeypatch):
    monkeypatch.setenv("FINANCEOS_MFA_ENCRYPTION_KEY", Fernet.generate_key().decode())
    engine = create_engine("sqlite://", connect_args={"check_same_thread": False}, poolclass=StaticPool)
    Base.metadata.create_all(engine)
    fabrica = sessionmaker(bind=engine)
    monkeypatch.setattr(auth_service, "get_session", fabrica)
    service = AuthService()
    usuario = service.registrar("Persona", "seguridad@financeos.local", "Clave-Mfa-Segura-2026")
    preparacion = service.preparar_mfa(usuario.id)
    service.confirmar_mfa(usuario.id, MfaService.codigo(preparacion["secreto"]))
    for _ in range(service.MAX_INTENTOS_MFA):
        with pytest.raises(ValueError):
            service.iniciar(usuario.correo, "Clave-Mfa-Segura-2026", "000000")
    assert service.db.get(type(usuario), usuario.id).bloqueado_hasta is not None

    usuario.bloqueado_hasta = None
    usuario.intentos_fallidos = 0
    service.db.commit()
    _, token = service.iniciar(usuario.correo, "Clave-Mfa-Segura-2026", MfaService.codigo(preparacion["secreto"]))
    sesion = service.db.query(SesionUsuario).filter_by(token_hash=service._hash_token(token)).one()
    sesion.ultima_actividad = datetime.now() - service.INACTIVIDAD_MAXIMA - timedelta(seconds=1)
    service.db.commit()
    assert service.autenticar(token) is None
    assert sesion.revocada_en is not None


def test_mfa_descifra_clave_anterior_y_recifra_con_actual(monkeypatch):
    anterior, actual = Fernet.generate_key().decode(), Fernet.generate_key().decode()
    monkeypatch.setenv("FINANCEOS_MFA_ENCRYPTION_KEY", anterior)
    monkeypatch.setenv("FINANCEOS_MFA_CURRENT_KEY_ID", "old")
    cifrado = MfaService().cifrar("SECRETO")
    monkeypatch.setenv("FINANCEOS_MFA_ENCRYPTION_KEY", actual)
    monkeypatch.setenv("FINANCEOS_MFA_CURRENT_KEY_ID", "new")
    monkeypatch.setenv("FINANCEOS_MFA_PREVIOUS_KEYS", f"old:{anterior}")
    servicio = MfaService()
    assert servicio.descifrar(cifrado) == "SECRETO"
    assert servicio.necesita_rotacion(cifrado)

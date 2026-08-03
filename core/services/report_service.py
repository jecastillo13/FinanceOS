import csv
from calendar import monthrange
from datetime import date
from io import BytesIO, StringIO

from sqlalchemy.orm import joinedload

from core.database import get_session
from core.models import Categoria, Movimiento
from core.services.exchange_service import ExchangeService


class ReportService:
    """Prepara reportes sin cargar exportadores pesados durante la navegación."""

    def __init__(self):
        self.db = get_session()
        self.exchange = ExchangeService()

    def obtener_reporte(self, anio, mes):
        inicio = date(anio, mes, 1)
        fin = date(anio, mes, monthrange(anio, mes)[1])
        movimientos = (
            self.db.query(Movimiento)
            .options(joinedload(Movimiento.cuenta), joinedload(Movimiento.categoria))
            .join(Categoria)
            .filter(
                Movimiento.fecha.between(inicio, fin),
                Categoria.tipo.in_(("Ingreso", "Gasto")),
            )
            .order_by(Movimiento.fecha.desc(), Movimiento.id.desc())
            .all()
        )

        filas = []
        monedas_sin_tasa = set()
        ingresos_cop = 0.0
        gastos_cop = 0.0
        for movimiento in movimientos:
            moneda = (movimiento.cuenta.moneda or "COP").upper()
            valor_cop = self.exchange.convertir(abs(movimiento.valor), moneda, "COP")
            if valor_cop is None:
                monedas_sin_tasa.add(moneda)
            elif movimiento.categoria.tipo == "Ingreso":
                ingresos_cop += valor_cop
            else:
                gastos_cop += valor_cop

            filas.append(
                {
                    "fecha": movimiento.fecha,
                    "tipo": movimiento.categoria.tipo,
                    "descripcion": movimiento.descripcion or "Sin descripción",
                    "categoria": movimiento.categoria.nombre,
                    "grupo": movimiento.categoria.grupo or "Otros",
                    "cuenta": movimiento.cuenta.nombre,
                    "moneda": moneda,
                    "valor_original": abs(movimiento.valor),
                    "valor_cop": valor_cop,
                    "observaciones": movimiento.observaciones or "",
                }
            )

        return {
            "anio": anio,
            "mes": mes,
            "filas": filas,
            "ingresos_cop": round(ingresos_cop, 2),
            "gastos_cop": round(gastos_cop, 2),
            "balance_cop": round(ingresos_cop - gastos_cop, 2),
            "monedas_sin_tasa": sorted(monedas_sin_tasa),
        }

    @staticmethod
    def generar_csv(reporte):
        salida = StringIO(newline="")
        escritor = csv.writer(salida)
        escritor.writerow(
            ["Fecha", "Tipo", "Descripción", "Categoría", "Grupo", "Cuenta", "Moneda", "Valor original", "Valor COP", "Observaciones"]
        )
        for fila in reporte["filas"]:
            escritor.writerow(
                [
                    fila["fecha"].isoformat(), fila["tipo"], fila["descripcion"], fila["categoria"],
                    fila["grupo"], fila["cuenta"], fila["moneda"], fila["valor_original"],
                    "" if fila["valor_cop"] is None else fila["valor_cop"], fila["observaciones"],
                ]
            )
        return salida.getvalue().encode("utf-8-sig")

    @staticmethod
    def generar_excel(reporte):
        # Importación diferida: abrir Reportes sigue siendo rápido.
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

        libro = Workbook()
        resumen = libro.active
        resumen.title = "Resumen"
        resumen.sheet_view.showGridLines = False
        resumen.merge_cells("A1:D1")
        resumen["A1"] = f"FinanceOS - Reporte mensual {reporte['anio']}-{reporte['mes']:02d}"
        resumen["A1"].font = Font(size=18, bold=True, color="FFFFFF")
        resumen["A1"].fill = PatternFill("solid", fgColor="172554")
        resumen["A1"].alignment = Alignment(horizontal="left")
        resumen.row_dimensions[1].height = 30
        resumen.append([])
        resumen.append(["Indicador", "Valor (COP)"])
        resumen.append(["Ingresos", reporte["ingresos_cop"]])
        resumen.append(["Gastos", reporte["gastos_cop"]])
        resumen.append(["Balance", reporte["balance_cop"]])

        movimientos = libro.create_sheet("Movimientos")
        movimientos.sheet_view.showGridLines = False
        encabezados = ["Fecha", "Tipo", "Descripción", "Categoría", "Grupo", "Cuenta", "Moneda", "Valor original", "Valor COP", "Observaciones"]
        movimientos.append(encabezados)
        for fila in reporte["filas"]:
            movimientos.append(
                [fila["fecha"], fila["tipo"], fila["descripcion"], fila["categoria"], fila["grupo"],
                 fila["cuenta"], fila["moneda"], fila["valor_original"], fila["valor_cop"], fila["observaciones"]]
            )

        azul = PatternFill("solid", fgColor="1E3A8A")
        borde = Border(bottom=Side(style="thin", color="CBD5E1"))
        for hoja in (resumen, movimientos):
            for celda in hoja[3] if hoja is resumen else hoja[1]:
                celda.fill = azul
                celda.font = Font(bold=True, color="FFFFFF")
                celda.alignment = Alignment(vertical="center")
            hoja.freeze_panes = "A4" if hoja is resumen else "A2"

        for fila in resumen.iter_rows(min_row=4, max_row=6, min_col=1, max_col=2):
            for celda in fila:
                celda.border = borde
            fila[1].number_format = '$#,##0.00;[Red]($#,##0.00);-'
        for fila in movimientos.iter_rows(min_row=2):
            fila[0].number_format = "yyyy-mm-dd"
            fila[7].number_format = '#,##0.00;[Red](#,##0.00);-'
            fila[8].number_format = '$#,##0.00;[Red]($#,##0.00);-'
            for celda in fila:
                celda.border = borde
                celda.alignment = Alignment(vertical="top", wrap_text=celda.column in (3, 10))
        movimientos.auto_filter.ref = movimientos.dimensions
        anchos = {"A": 13, "B": 12, "C": 30, "D": 20, "E": 18, "F": 20, "G": 11, "H": 16, "I": 16, "J": 34}
        for columna, ancho in anchos.items():
            movimientos.column_dimensions[columna].width = ancho
        resumen.column_dimensions["A"].width = 24
        resumen.column_dimensions["B"].width = 20

        salida = BytesIO()
        libro.save(salida)
        return salida.getvalue()

    @staticmethod
    def generar_pdf(reporte):
        # Importación diferida para no penalizar el arranque de Streamlit/API.
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.lib.units import mm
        from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

        salida = BytesIO()
        documento = SimpleDocTemplate(
            salida, pagesize=landscape(A4), rightMargin=14 * mm, leftMargin=14 * mm,
            topMargin=14 * mm, bottomMargin=14 * mm,
        )
        estilos = getSampleStyleSheet()
        elementos = [
            Paragraph(f"FinanceOS - Reporte mensual {reporte['anio']}-{reporte['mes']:02d}", estilos["Title"]),
            Spacer(1, 5 * mm),
        ]
        resumen = Table(
            [["Ingresos (COP)", "Gastos (COP)", "Balance (COP)"],
             [f"{reporte['ingresos_cop']:,.2f}", f"{reporte['gastos_cop']:,.2f}", f"{reporte['balance_cop']:,.2f}"]],
            colWidths=[80 * mm] * 3,
        )
        resumen.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1E3A8A")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTNAME", (0, 1), (-1, 1), "Helvetica-Bold"),
            ("FONTSIZE", (0, 1), (-1, 1), 14),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("BOX", (0, 0), (-1, -1), 0.5, colors.HexColor("#CBD5E1")),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
            ("TOPPADDING", (0, 0), (-1, -1), 8),
        ]))
        elementos.extend([resumen, Spacer(1, 7 * mm)])

        datos = [["Fecha", "Tipo", "Descripción", "Categoría", "Cuenta", "Moneda", "Valor", "Valor COP"]]
        for fila in reporte["filas"]:
            datos.append([
                fila["fecha"].strftime("%d/%m/%Y"), fila["tipo"], fila["descripcion"], fila["categoria"],
                fila["cuenta"], fila["moneda"], f"{fila['valor_original']:,.2f}",
                "Sin tasa" if fila["valor_cop"] is None else f"{fila['valor_cop']:,.2f}",
            ])
        if len(datos) == 1:
            datos.append(["-", "-", "No hay movimientos en este período", "-", "-", "-", "-", "-"])
        tabla = Table(datos, repeatRows=1, colWidths=[23*mm, 20*mm, 53*mm, 37*mm, 35*mm, 17*mm, 28*mm, 30*mm])
        tabla.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#172554")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("ALIGN", (-3, 1), (-1, -1), "RIGHT"),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F1F5F9")]),
            ("LINEBELOW", (0, 0), (-1, -1), 0.25, colors.HexColor("#CBD5E1")),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
            ("TOPPADDING", (0, 0), (-1, -1), 5),
        ]))
        elementos.append(tabla)
        documento.build(elementos)
        return salida.getvalue()

    def cerrar(self):
        self.exchange.cerrar()
        self.db.close()
